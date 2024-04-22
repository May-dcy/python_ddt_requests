from utils.handle_path import *
import os
import openpyxl
'''
pip install openpyxl    下载openpyxl库
此模块是用来读取表格的工具类
'''
'''封装一个班读取Excel表格的工具类'''
class Handle_Excel(object):

    def __init__(self,filename,sheet_name):
        self.filename=filename#把传进来的形式参数赋值给到实例变量self.filename
        self.sheet_name=sheet_name

    '''封装一个打开Excel表格的工具方法'''
    def open(self):
        # 通过openpyxl模块调用load_workbook函数打开filename文件
        self.wb=openpyxl.load_workbook(self.filename)
        # 通过self.wb这个Excel文件的对象读取对应的sheet页面的
        self.sh=self.wb[self.sheet_name]

    '''封装了一个读取Excel表格的工具方法'''
    def read_data(self):
        self.open()#打开Excel表格
        # 获取表格中的每一行的数据放在元组当中
        datas=list(self.sh.rows)#把每一行的元素数据放在list列表当中
        # print(datas)
        title = [i.value for i in datas[0]]  # 对这个datas进行for循环遍历并且赋值给title这个变量
        # print(title)
        # 创建一个空列表，用来接收所有的测试用例
        cases=[]
        for j in datas[1:]:#对第2,3,4,5,6......行进行遍历
            # print(j)
            values=[k.value for k in j]#然后通过对j进行遍历，用k.value获取没有用例对象中value值，然后放在一个列表
            # print(values)
            case=dict(zip(title,values))#把title列表和values每个用例一一对应打包放进一个字典当中
            # print(case)
            cases.append(case)
        return  cases # 把所有的用例返回
    '''
    封装一个往Excel表格里面写入测试结果的工具方法
    '''
    def write_excel(self,row,column,value=None):

        self.open()# 打开表格
        self.sh.cell(row, column, value)# 往固定的row行和column列写入value数据
        self.wb.save(self.filename)# 保存数据

# if __name__ == '__main__':
#     path=os.path.join(data_path,"data.xlsx")
#     read_excel=Handle_Excel(path,"login")
#     read_excel.write_excel(2,8,"通过")


